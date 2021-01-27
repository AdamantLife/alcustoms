## Test Target
from alcustoms.excel import Tables
## Test Framework
import unittest
## Testing utilities
from alcustoms.excel import tests
## Super Module
import openpyxl.worksheet.table
## This Module
from alcustoms import excel
from alcustoms.excel.Ranges import tuple_to_range

class TableTests(unittest.TestCase):
    def setUp(self):
        tests.basicsetup(self)
        return super().setUp()

    def test_get_all_tables(self):
        tables = Tables.get_all_tables(self.workbook)
        self.assertEqual(len(tables),3)
        tablelookup = {table.name:{"worksheet":worksheet,"table":table} for worksheet,table in tables}
        for tablename,table in tests.DATA['TABLES'].items():
            with self.subTest(tablename=tablename,table = table):
                worksheet = getattr(self,table['worksheet'])
                ## Check correct name
                self.assertIn(table["name"],tablelookup)
                lookup = tablelookup[table["name"]]
                ## Check correct type
                self.assertIsInstance(lookup["table"],Tables.EnhancedTable)
                ## Check correct worksheet
                self.assertEqual(lookup["worksheet"],worksheet)
                ## Check correct range
                self.assertEqual(lookup["table"].ref,table["range"])


    def test_get_all_tables_ramifications(self):
        """ Initiating a Table seems to automatically add the table to the parent-worksheet's _tables attribute. So we're going to test that by doing a get_all_tables which converts tables """
        for worksheetname in self.workbook.sheetnames:
            worksheet = self.workbook[worksheetname]
            with self.subTest(worksheet = worksheet):
                ## New version of worksheet._tables is a dict subclass
                if isinstance(worksheet._tables, dict): tables = list(worksheet._tables.values())
                else: tables = list(worksheet._tables)

                ##                               All Tables are Tables                          but not EnhancedTable (Subclass)
                self.assertTrue(all(isinstance(table,openpyxl.worksheet.table.Table) and not isinstance(table,Tables.EnhancedTable) for table in tables))

        Tables.get_all_tables(self.workbook)

        for worksheetname in self.workbook.sheetnames:
            worksheet = self.workbook[worksheetname]
            with self.subTest(worksheet = worksheet):
                ## New version of worksheet._tables is a dict subclass
                if isinstance(worksheet._tables, dict): tables = list(worksheet._tables.values())
                else: tables = list(worksheet._tables)

                self.assertTrue(all(isinstance(table,Tables.EnhancedTable) for table in tables))

class EnhancedTableTests(unittest.TestCase):
    def setUp(self):
        tests.basicsetup(self)
        self.tables = {table.name:table for worksheet,table in Tables.get_all_tables(self.workbook)}
        return super().setUp()

    def test_enhancetable(self):
        """ Tests attributes of the EnhancedTable, ensuring that inherited attributes exist and new methods function as intended """
        ## Technically, enchancedTable.worksheet was tested in get_all_tables
        for tablename,table in tests.DATA['TABLES'].items():
            with self.subTest(tablename=tablename,table = table):
                worksheet = getattr(self,table['worksheet'])
                enhancedtable = self.tables[table['name']]

                ## Check to make sure everything is here
                ## Inheritted Stuff
                for attr in [
                    ## Attribtues
                    "id", "displayName", "name", "comment", "ref", "tableType", "headerRowCount",
                    "insertRow", "insertRowShift", "totalsRowCount", "totalsRowShown", "published",
                    "headerRowDxfId", "dataDxfId", "totalsRowDxfId", "headerRowBorderDxfId", "tableBorderDxfId",
                    "totalsRowBorderDxfId", "headerRowCellStyle", "dataCellStyle", "totalsRowCellStyle",
                    "connectionId", "autoFilter", "sortState", "tableColumns", "tableStyleInfo",
                    ## Methods
                    "to_tree","path","_write","_initialise_columns"
                    ]:
                    with self.subTest(enhancedtable=enhancedtable,attr=attr):
                        self.assertTrue(hasattr(enhancedtable,attr))

                ## New Stuff
                for attr in ["worksheet","range","headers","todicts"]:
                    with self.subTest(enhancedtable=enhancedtable, attr=attr):
                        self.assertTrue(hasattr(enhancedtable,attr))

                ## Individual tests
                self.assertEqual(enhancedtable.range,   excel.Range(worksheet,table['range']))
                self.assertEqual(enhancedtable.headerrange(), excel.Range(worksheet,table['headerrange']))
                self.assertEqual(enhancedtable.headers(), table['headers'])
                self.assertEqual(enhancedtable.datarange(), excel.Range(worksheet,table['datarange']))
                self.assertEqual(enhancedtable.todicts(),[table['headers'],]+[dict(list(zip(table['headers'],row))) for row in table['values']])        

    def test_insertcolumn(self):
        """ Tests that enhanced table can insert a column and update it's reference """

class MethodCase(unittest.TestCase):
    """ TestCase for various utility methods """
    def test_tuple_to_range(self):
        """ Tests various values for tuple_to_range """
        for test,absolute,result in [(  (1,1,1,1),  False,  "A1:A1" ),
                                     (  (1,1,1,1),  True,  "$A$1:$A$1" ),
                                     (  ((1,1,1,1),),  False,  "A1:A1" ),
                                     (  ((1,1,1,1),),  True,  "$A$1:$A$1" ),
                            ]:
            with self.subTest(test = test, result = result, absolute = absolute):
                self.assertEqual(tuple_to_range(*test, absolute = absolute),result)


    def test_gettablesize(self):
        """ Tests gettablesize for a variety of table shapes """
        tests.basicsetup(self)
        sheet = self.sheet_Tables3
        r1,c1 = 2,3
        self.assertEqual(Tables.gettablesize(sheet,c1,r1),"C2:F15")

    def test_gettablesize_blank_columns(self):
        """ Tests gettablesize with blank columns """
        tests.basicsetup(self)
        sheet = self.sheet_Tables4
        r1, c1 = 1,1
        for (greedycolumns,result) in [(1,"A1:C3"),(2, "A1:G3")]:
            with self.subTest(greedycolumns = greedycolumns, result = result):
                self.assertEqual(Tables.gettablesize(sheet, c1, r1, greedycolumns=greedycolumns), result)

    def test_gettablesize_blank_rows(self):
        """ Tests gettablesize with blank rows """
        tests.basicsetup(self)
        sheet = self.sheet_Tables4
        r1, c1 = 1,11
        for (greedyrows,result) in [(1,"K1:N4"),(2, "K1:N7")]:
            with self.subTest(greedyrows = greedyrows, result = result):
                self.assertEqual(Tables.gettablesize(sheet, c1, r1, greedyrows=greedyrows), result)

    def test_gettablesize_notable(self):
        """ Tests that gettablesize on a blank row/column results in None"""
        tests.basicsetup(self)
        ## This uses the blank space between the two greedy table tests
        sheet = self.sheet_Tables4
        r1, c1 = 1,10
        self.assertIsNone(Tables.gettablesize(sheet, c1, r1))

    def test_gettablesize_singlecolumn(self):
        """ Tests that single column tables work """
        tests.basicsetup(self)
        ## This uses the first column of the greedycolumn table
        sheet = self.sheet_Tables4
        r1,c1 = 1,1
        self.assertEqual(Tables.gettablesize(sheet, c1, r1), "A1:A3")


    def test_get_table_by_name(self):
        """ Basic Tests for get_table_by_name """
        tests.basicsetup(self)
        sheet = self.sheet_Tables1
        name = 'testtable_1'
        table = Tables.get_table_by_name(sheet,name)
        self.assertIsInstance(table,openpyxl.worksheet.table.Table)
        import collections
        self.assertEqual(dict(a = 1),collections.OrderedDict([("a",1),]))
        dicts = Tables.EnhancedTable.from_table(table,sheet).todicts()
        self.assertEqual(Tables.EnhancedTable.from_table(table,sheet).todicts(),[["Name","Value"],dict(Name="Hello",Value = 1),dict(Name="World",Value=2)])

    def test_dicts_to_table_simple_add(self):
        """ A simple test to add a new table based on dicts to a worksheet """
        tests.basicsetup(self)
        ws = self.workbook.create_sheet("testsheet")
        testdata = [dict(a = 1, b = 1, c = 1),
                    dict(a = 2, b = 2, c = 2),
                    dict(a = 3, b = 3, c = 3)]
        table = Tables.dicts_to_table(ws, testdata, start = "$A$1")
        testrange = excel.Range(ws,Tables.gettablesize(ws,1,1))
        self.assertEqual(testrange.startcoord,excel.Coordinate("A1"))
        self.assertEqual(testrange.endcoord,excel.Coordinate("C4"))
        output = table.todicts()[1:]
        self.assertEqual(len(output),len(testdata))
        self.assertEqual(output,testdata)

if __name__ == "__main__":
    unittest.main()