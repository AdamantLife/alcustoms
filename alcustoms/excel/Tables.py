""" alcustoms.excel.Tables

    Provides an "EnhancedTable" Subclass class which extends the functionality of normal Tables
"""
## Super Module
from openpyxl.worksheet.table import Table, TableColumn
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
## This Module
from alcustoms.excel import Coordinate, Range
from alcustoms.excel.Ranges import tuple_to_range
## Builtin
import collections
import itertools

__all__ = ["EnhancedTable",]

class EnhancedTable(Table):
    """ A better Table Class, returned by get_all_tables """
    def from_table(table,worksheet):
        def oldversion():
            ## NOTE! It seems that Initializing a Table object (which EnhancedTable is a subclass of) automatically adds the table to the Worksheet's _tables list
            ## This means that we have to check and remove the original Table object, and should return any pre-built EnhancedTable we find.
            oldtables = [tab for tab in worksheet._tables if tab == table]
            for tab in oldtables:
                ## Don't bother making a new one
                if isinstance(tab,EnhancedTable): return tab
                ## Remove the old-style one
                else:
                    worksheet._tables.remove(tab)
        def newversion():
            if table.name in worksheet._tables:
                if isinstance((ntable := worksheet._tables[table.name]), EnhancedTable):
                    return ntable
                del worksheet._tables[table.name]

        ## New version of openpyxl changes worksheet._tables to a dict subclass
        if isinstance(worksheet._tables, dict): newversion()
        else: oldversion()
        ## If no EnhancedTable version found, create one (which- again- seems to automatically be added to the _tables list)
        return EnhancedTable(worksheet = worksheet,
                             id=table.id, displayName=table.displayName, ref=table.ref, name=table.name, comment=table.comment, tableType=table.tableType,
                             headerRowCount=table.headerRowCount, insertRow=table.insertRow, insertRowShift=table.insertRowShift, totalsRowCount=table.totalsRowCount, totalsRowShown=table.totalsRowShown,
                             published=table.published, headerRowDxfId=table.headerRowDxfId, dataDxfId=table.dataDxfId, totalsRowDxfId=table.totalsRowDxfId, headerRowBorderDxfId=table.headerRowBorderDxfId,
                             tableBorderDxfId=table.tableBorderDxfId, totalsRowBorderDxfId=table.totalsRowBorderDxfId, headerRowCellStyle=table.headerRowCellStyle, dataCellStyle=table.dataCellStyle,
                             totalsRowCellStyle=table.totalsRowCellStyle, connectionId=table.connectionId, autoFilter=table.autoFilter, sortState=table.sortState, tableColumns=table.tableColumns,
                             tableStyleInfo=table.tableStyleInfo, extLst=table.extLst)
    def __init__(self,worksheet, **kw):
        self.worksheet = None
        super().__init__(**kw)
        self.worksheet = worksheet
        self.range = Range(worksheet,self.ref)
        ## For the moment, we are disabling this functionality
        ## The serializer on Table seems to be serializing other attributes
        ## self.worksheet.add_table(self)

    @property
    def ref(self):
        return self.range.range
    @ref.setter
    def ref(self,value):
        self.range = Range(self.worksheet,value)

    def insertcolumn(self,index = None,columnname = None):
        collen = len(self.tableColumns)
        if index is None:
            index = collen - 1
        if index >= collen: raise ValueError("Index outside of Table Range")
        if columnname is None:
            columnname = f"Column{collen}"
        if columname in [column.name for column in self.tableColumns]:
            columnname = columnname + " 2"
        self.tableColumns.insert(index,TableColumn(id=collen,name=columnname))
        self.range += [1,None]

    def headers(self, attribute = "value"):
        """ Returns the headers of the EnhancedTable as a list, via self.range.row """
        ## Last row (equal to row count) should contain actual headers... In theory
        ## And we zero-index it
        headerrow = self.headerRowCount - 1
        return list(self.range.row(headerrow,attribute=attribute))

    def addheaders(self, headers):
        """ Adds a set of headers to the table at the bottom of the header range """
        headerlen = len(headers)               

    def headerrange(self):
        """ Returns the Table's headers as a Range object """
        headerlength = self.headerRowCount
        columnlength = len(self.tableColumns)
        return self.range.subrange(None,(str(headerlength-1),str(columnlength-1)))

    def datarange(self):
        """ Returns the Table's body as a Range object """
        headerlength = self.headerRowCount
        columnlength = len(self.tableColumns)
        return self.range.subrange(
            (str(headerlength),str(0)),None)

    def todicts(self,keyfactory = None):
        """ Converts all data rows to dicts based on column headers. The first element of the returned list is a list of the header strings used.
        
        keyfactory is an callback function to modify the keys (example- the lowerstrip lambda available in this module executes
        key.lower().replace(" ","_") to lowercase and remove spaces in keys).
        """
        if keyfactory is None: keyfactory = lambda key: key
        headers = [keyfactory(key) for key in self.headers()]

        data = [collections.OrderedDict(list(zip(headers,row))) for row in self.datarange().rows_from_range()]
        data.insert(0,headers)
        return data

def get_all_tables(workbook):
    """ Returns a list of tuples of all tables in the workbook formatted: (worksheetobject, EnhancedTable Object) """
    out = []
    for worksheetname in workbook.sheetnames:
        worksheet = workbook[worksheetname]
        ## To ensure list integrity, we'll have to copy the list
        ## (Initiating Tables seems to automatically add them to _tables list)
        tables = list(worksheet._tables)
        for table in tables:
            ## New version of openpyxl changes worksheet._tables to a dict subclass
            if isinstance(table, str):
                table = worksheet._tables[table]
            ## Check if it was pre-converted and return it if so
            if isinstance(table,EnhancedTable):
                out.append((worksheet,table))
            ## Otherwise, convert it ourselves
            else:
                out.append((worksheet,EnhancedTable.from_table(table,worksheet)))
    return out

def get_table_by_name(source,name):
    """ Returns the table with the given displayName.

        source should be a Worksheet or Workbook.
        name should be the displayName.
    
        Workbooks that have duplicate tables are considered Invalid by Excel,
        so if this method finds multiple tables with the given displayName it
        will raise a ValueError.
    """
    if not isinstance(source,(Workbook,Worksheet)):
        raise TypeError("source should be a Workbook or Worksheet object")
    if not isinstance(name,str):
        raise TypeError("name should be a string")
    if isinstance(source, Workbook):
        sheets = source.worksheets
    else:
        sheets = [source,]
    results = []
    for worksheet in sheets:
        ## New version of openpyxl changes ._tables to a dict subclass
        if isinstance(worksheet._tables, dict):
            if name in worksheet._tables:
                results.append((worksheet, worksheet._tables[name]))
        ## Old version was a list
        else:
            for table in worksheet._tables:
                if table.displayName == name:
                    results.append((worksheet,table))
    if len(results) == 0:
        return None
    if len(results) > 1:
        raise ValueError(f'Got multiple values for "{name}"')
    sheet,table = results[0]
    if isinstance(table,EnhancedTable): return table
    return EnhancedTable.from_table(table,sheet)

def dicts_to_table(sheet, dinput, tablename = None, start = None, headers = None):
    """ Writes a list of dictionaries or lists into a Table.
  
        sheet should be a Worksheet.
        dinput should be a list of either lists or dictionaries.
        If provided, tablename should be a string. If the tablename is defined
        within the worksheet, then all rows in the given table will be replaced
        with dinput.
        If provided, start should be a valid coordinate to place the table. It is
        a SyntaxError to supply both an existing tablename and start and a ValueError
        to supply neither.
        headers should be a list of strings. headers is optional if dinput is a
        list of dicts. If so, then only the given keys will be outputted from dinput.
        If dinput is a list and tablename is not an existing table, then headers is
        required. Otherwise (dinput is a list and tablename exists), then headers is
        ignored.
    """
    if not isinstance(sheet,Worksheet):
        raise TypeError('sheet must be a worksheet')
    if not isinstance(dinput,(list,tuple)) or not all(isinstance(item,(dict,list,tuple)) for item in dinput):
        raise TypeError('dinput should be a list of lists or dicts')
    t1 = dinput[0].__class__
    if any(not isinstance(item,t1) for item in dinput):
        raise TypeError("All dinputs must be the same class")
    if tablename is None: tablename = ""
    if not isinstance(tablename,str):
        raise TypeError('If supplied, tablename should be a string')
    table = None
    if tablename:
        table = get_table_by_name(sheet,tablename)
    else:
        ## AutoGenerate Tablename
        i = 1
        while not tablename:
            tablename = f"Table{i}"
            t = get_table_by_name(sheet,tablename)
            if t: tablename = None

    if start:
        start = Coordinate(start)
        if table and table.startcoord != start:
            raise SyntaxError("Cannot supply both an existing table and start")
    elif table:
        start = table.startcoord
    ## Not start and not table
    else:
        raise ValueError("dicts_to_table requires either start or an existing tablename")

    if headers:
        if not isinstance(headers,(list,tuple)) or any(not isinstance(item,str) for item in headers):
            raise ValueError("headers should be a string")
        ## Ignore headers if existing table
        if table: headers = None

    ## Normalize Output
    if headers is None:
        if table:
            headers = table.headers()

    ## Normalize dinput structure
    if issubclass(t1,dict):
        if headers is None:
            iheaders = list(itertools.chain.from_iterable([item.keys() for item in dinput]))
            ## Instead of using set, we're going to attempt to conserve at least some of the order of the keys
            ## (at time of writing, all standard dicts memorize input order of keys, which makes them reliable)
            headers = []
            for h in iheaders:
                if h not in headers: headers.append(h)
        dinput = [[item.get(key,"") for key in headers] for item in dinput]
    ## Otherwise, dinput items are lists or tuples (or- theoretically- subclasses), so don't do anything
    
    ## If existing table, remove it and rewrite it
    if table:
        for cell in table.range.cells_by_row(value = "cell"):
            cell.value = None
        table.worksheet._tables.remove(table)
    ## TODO: we should probably make some attempt to NOT overwrite data in proximity if table extends beyond original constraints

    ## Write Table Headers
    for coffset,column in enumerate(headers):
        offset = Coordinate(row=str(0),column=str(coffset))
        cell = start + offset
        cell = sheet[cell.toA1string()]
        cell.value = str(column)
    ## Write table data
    ## Start writing row at 1
    for roffset,row in enumerate(dinput, start = 1):
        for coffset,column in enumerate(headers):
            offset = Coordinate(str(roffset),str(coffset))
            cell = start + offset
            cell = sheet[cell.toA1string()]
            cell.value = row[coffset]
    ## Add/Re-add Table to sheet
    endcell = Coordinate(str(roffset),str(coffset))
    endcell = start + endcell
    table = Table(displayName = tablename, ref = f"{start.toA1string(False)}:{endcell.toA1string(False)}")
    sheet.add_table(table)
    table = EnhancedTable.from_table(table,sheet)
    return table

def gettablesize(sheet,startcolumn,startrow, absolute = False, greedycolumns = False, greedyrows = False):
    """ This is a function to intuit the shape of a data series which is laid out in a table format.

        It works by assuming the top row to be the header row. It scans this row until it reaches a blank cell.
        It then iterates over the subsequent rows up to the end of the header row. When a row is completely
        blank it breaks and returns a rage from the first cell of the header row to the last row and last column
        of the header row.

        The absolute keyword determines whether the returned range is absolute (default False).

        greedycolumns and greedyrows allow for the given number of blank, consecutive columns or rows (respectively) to be
        skipped when determining the table size.

        e.x.- Table Starting at (1,1) [A1] with no greedy columns or rows
        ---------------
        |W   X   Y   Z|
        |1   2   3    |
        |    4       5|
        |6            |
        |             |
        |[... etc ...]|
        ---------------
        Returns a range "A$1$:D$4$"
    """
    row,column = startrow,startcolumn
    cell = sheet.cell(row = row, column = column)
    blank = int(not bool(cell.value))
    ## If cell is blank is not blank under any circumstances it should be counted
    ## If greedycolumns then continue on the first blank if current cell is blank
    while not blank or (greedycolumns and blank <= greedycolumns):
        column += 1
        cell = sheet.cell(row = row, column = column)
        if cell.value: blank = 0
        else: blank +=1 

    endcolumn = column - blank 
    if endcolumn < startcolumn:
        return None

    ## Check row 
    def blankrow(row):
        col = 1
        while col <= endcolumn:
            if sheet.cell(row = row, column = col).value:
                return False
            col += 1
        return True
    ## Start check table data 
    row = 2 
    norow = int(blankrow(row))
    ## See notes above on greedycolumns
    while not norow or (greedyrows and norow <= greedyrows):
        row += 1
        if blankrow(row): norow += 1
        else: norow = 0
    ## Current Row does not have value
    endrow = row - norow
    return tuple_to_range(startcolumn,startrow,endcolumn,endrow, absolute = absolute)