## Tested Module
from alcustoms import excel
import openpyxl.worksheet.table
## Builtin
import json
import pathlib
import types
import unittest

DIRECTORY = pathlib.Path(__file__).resolve().parent
DATAFILE = (DIRECTORY / "exceltests.json").resolve()
TESTFILE = (DIRECTORY / "testbook.xlsx").resolve()

def loaddata():
    """ Loads testdata from DATAFILE """
    with open(DATAFILE,'r') as f:
        data = json.load(f)
    ## Time-saving for tests, space-saving for jsonfile, avoids clerical errors
    for _range in data['RANGES'].values():
        _range['column_values'] = list(zip(*_range['row_values']))
    return data

DATA = loaddata()

def getobjectreference(objecttype):
    """ Searches the global scope and the excel module for the provided reference string """
    try: factory = globals[objecttype]
    except KeyError:
        factory = getattr("excel",objecttype)
        if not factory: raise ValueError(f"Test could not build: {objecttype}")
    return factory

def lambdafactory(*args,**kw):
    """ A callable to create a lambda for testing """
    return lambda *args, **kwargs: None

class TestObject():
    """ A simple object for testing """
    pass

def basicsetup(testcase):
    """ Does basic workbook setup for the testcase """
    testcase.workbook = excel.load_workbook(str(TESTFILE), data_only = True)
    for sheet in DATA['SHEETS']:
        setattr(testcase,sheet['alias'],testcase.workbook[sheet['name']])


class CoordinateTests(unittest.TestCase):
    def setUp(self):
        basicsetup(self)
        return super().setUp()
    
    def test_creation(self):
        """ Test the creation of a Coordinate via various methods and validity of data """
        for coordinate in DATA['COORDINATES']:
            with self.subTest(coordinate = coordinate):
                result = coordinate.get('result')
                if not result: continue
                definition = coordinate.get('definition')
                objecttype = coordinate.get("object")
                if objecttype:
                    factory = getobjectreference(objecttype)
                    definition = factory(definition)

                ## Tests
                if result['type'] == "success":
                    if isinstance(definition,(tuple,list)):
                        coord = excel.Coordinate(*definition)
                    else:
                        coord= excel.Coordinate(definition)
                    self.assertEqual(coord._row.value,result['row']['value'])
                    self.assertEqual(coord._row.absolute,result['row']['absolute'])
                    self.assertEqual(coord._column.value,result['column']['value'])
                    self.assertEqual(coord._column.absolute,result['column']['absolute'])
                elif result['type'] == "exception":
                    error = getattr(__builtins__,result['e_type'])

                    ## Added length check to handle different length objects
                    ## (that are expected to fail further downstream)
                    if isinstance(definition,(tuple,list)) and 0 < len(definition) < 3:
                        self.assertRaisesRegex(error,result['e_regex'],
                                               excel.Coordinate,*definition)
                    elif isinstance(definition,dict):
                        self.assertRaisesRegex(error,result['e_regex'],
                                               excel.Coordinate,**definition)
                    else:
                        self.assertRaisesRegex(error,result['e_regex'],
                                               excel.Coordinate,definition)

class RangeTests(unittest.TestCase):
    def setUp(self):
        basicsetup(self)
        return super().setUp()

    def test_range__eq__(self):
        """ Tests various equivalencies between Range Objects"""
        wb = self.workbook
        cs = self.sheet_CellNumbers
        testrange = excel.Range(cs,"$A$1:$J$1")
        testrange2 = excel.Range(cs,"$A$1:$J$1")
        testrange3 = excel.NamedRange(wb,"testrange_1").ranges[0]
        self.assertEqual(testrange,testrange2)
        self.assertEqual(testrange,testrange3)

    def test_testranges(self):
        """ Make sure testranges line up with what is actually in the Excel file """
        for solutionname,solution in DATA['RANGES'].items():
            with self.subTest(solutionname,solution = solution):
                testrange = excel.NamedRange(self.workbook,solution['name']).ranges[0]

                self.assertEqual([list(row) for row in testrange.rows_from_range("address")],solution['row_values'])
                self.assertEqual([list(row) for row in testrange.rows_from_range("value")],solution['row_values'])
                self.assertEqual([list(row) for row in testrange.rows_from_range("cell")],[list(self.sheet_CellNumbers[cell] for cell in row) for row in solution['row_values']])
        
    def test_rows_from_range(self):
        for solutionname,solution in DATA['RANGES'].items():
            with self.subTest(solutionname=solutionname, solution = solution):
                testrange = excel.NamedRange(self.workbook,solution['name']).ranges[0]
                rows = testrange.rows_from_range("cell")
                ## Assert Generator
                self.assertIsInstance(rows,types.GeneratorType)
                rows = list(rows)
                ## Assert Number of Rows
                self.assertEqual(len(rows),len(solution['row_values']))
                ## Assert length of row
                ## Note: casting is due to generators and mapping being used
                self.assertEqual(len(list(list(rows)[0])),
                                 len(solution['row_values'][0]))
                ## Assert Value is Default
                self.assertEqual(list(list(row) for row in testrange.rows_from_range()),list(list(row) for row in testrange.rows_from_range("value")))
                ## Assert Addresses
                self.assertEqual(list(list(row) for row in testrange.rows_from_range("address")), solution['row_values'])
                ## Assert Values
                self.assertEqual(list(list(row) for row in testrange.rows_from_range("value")),solution['row_values'])
                ## Assert Cells-type
                self.assertTrue(all(
                    all(isinstance(cell,excel.cell.Cell) for cell in row) for row in rows
                    ))


class TableTests(unittest.TestCase):
    def setUp(self):
        basicsetup(self)
        return super().setUp()

    def test_get_all_tables(self):
        tables = excel.get_all_tables(self.workbook)
        self.assertEqual(len(tables),3)
        tablelookup = {table.name:{"worksheet":worksheet,"table":table} for worksheet,table in tables}
        for tablename,table in DATA['TABLES'].items():
            with self.subTest(tablename=tablename,table = table):
                worksheet = getattr(self,table['worksheet'])
                ## Check correct name
                self.assertIn(table["name"],tablelookup)
                lookup = tablelookup[table["name"]]
                ## Check correct type
                self.assertIsInstance(lookup["table"],excel.EnhancedTable)
                ## Check correct worksheet
                self.assertEqual(lookup["worksheet"],worksheet)
                ## Check correct range
                self.assertEqual(lookup["table"].ref,table["range"])


    def test_get_all_tables_ramifications(self):
        """ Initiating a Table seems to automatically add the table to the parent-worksheet's _tables attribute. So we're going to test that by doing a get_all_tables which converts tables """
        for worksheetname in self.workbook.sheetnames:
            worksheet = self.workbook[worksheetname]
            with self.subTest(worksheet = worksheet):
                ##                               All Tables are Tables                          but not EnhancedTable (Subclass)
                self.assertTrue(all(isinstance(table,openpyxl.worksheet.table.Table) and not isinstance(table,excel.EnhancedTable) for table in worksheet._tables))

        excel.get_all_tables(self.workbook)

        for worksheetname in self.workbook.sheetnames:
            worksheet = self.workbook[worksheetname]
            with self.subTest(worksheet = worksheet):
                self.assertTrue(all(isinstance(table,excel.EnhancedTable) for table in worksheet._tables))

class EnhancedTableTests(unittest.TestCase):
    def setUp(self):
        basicsetup(self)
        self.tables = {table.name:table for worksheet,table in excel.get_all_tables(self.workbook)}
        return super().setUp()

    def test_enhancetable(self):
        """ Tests attributes of the EnhancedTable, ensuring that inherited attributes exist and new methods function as intended """
        ## Technically, enchancedTable.worksheet was tested in get_all_tables
        for tablename,table in DATA['TABLES'].items():
            with self.subTest(tablename=tablename,table = table):
                worksheet = getattr(self,table['worksheet'])
                enhancedtable = self.tables[table['name']]

                ## Check to make sure everything is here
                ## Inheritted Stuff
                for attr in [
                    ## Attribtues
                    "id", "displayName", "name", "comment", "ref", "tableType", "headerRowCount",
                    "insertRow", "insertRowShift", "totalsRowCount", "totalsRowShown", "published",
                    "headerRowDxfId", "dataDxfId", "totalsRowDxfId", "headerRowBorderDxfId", "tableBorderDxfId",
                    "totalsRowBorderDxfId", "headerRowCellStyle", "dataCellStyle", "totalsRowCellStyle",
                    "connectionId", "autoFilter", "sortState", "tableColumns", "tableStyleInfo",
                    ## Methods
                    "to_tree","path","_write","_initialise_columns"
                    ]:
                    with self.subTest(enhancedtable=enhancedtable,attr=attr):
                        self.assertTrue(hasattr(enhancedtable,attr))

                ## New Stuff
                for attr in ["worksheet","range","headers","todicts"]:
                    with self.subTest(enhancedtable=enhancedtable, attr=attr):
                        self.assertTrue(hasattr(enhancedtable,attr))

                ## Individual tests
                self.assertEqual(enhancedtable.range,   excel.Range(worksheet,table['range']))
                self.assertEqual(enhancedtable.headerrange(), excel.Range(worksheet,table['headerrange']))
                self.assertEqual(enhancedtable.headers(), table['headers'])
                self.assertEqual(enhancedtable.datarange(), excel.Range(worksheet,table['datarange']))
                self.assertEqual(enhancedtable.todicts(),[table['headers'],]+[dict(list(zip(table['headers'],row))) for row in table['values']])        

    def test_insertcolumn(self):
        """ Tests that enhanced table can insert a column and update it's reference """
        
class MethodCase(unittest.TestCase):
    """ TestCase for various methods """
    def test_tuple_to_range(self):
        """ Tests various values for tuple_to_range """
        for test,absolute,result in [(  (1,1,1,1),  False,  "A1:A1" ),
                                     (  (1,1,1,1),  True,  "$A$1:$A$1" ),
                                     (  ((1,1,1,1),),  False,  "A1:A1" ),
                                     (  ((1,1,1,1),),  True,  "$A$1:$A$1" ),
                            ]:
            with self.subTest(test = test, result = result, absolute = absolute):
                self.assertEqual(excel.tuple_to_range(*test, absolute = absolute),result)


    def test_gettablesize(self):
        """ Tests gettablesize for a variety of table shapes """
        basicsetup(self)
        sheet = self.sheet_Tables3
        r1,c1 = 2,3
        self.assertEqual(excel.gettablesize(sheet,c1,r1),"C2:F15")


if __name__ == "__main__":
    unittest.main()