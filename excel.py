## Builtin
import collections
import itertools
import re
## Third Party
from openpyxl import *
import openpyxl.worksheet.table, openpyxl.worksheet.worksheet
from openpyxl import utils

## A keyfactory for removing spaces and lowercasing Headers for EnhancedTable.todicts()
lowerstrip = lambda key: key.lower().strip().replace(" ","_")

class NamedRange():
    def __init__(self,workbook,name):
        self.workbook = workbook
        self.name = name
        self.ranges = [Range(workbook[worksheet],_range) for worksheet,_range in workbook.defined_names[name].destinations]

    def allcells(self, sort = "row"):
        """ Returns a flat list of all cells in the named range.

        if sort is "row" (default), the cells will be returned: first, in order
        of subrange; secondly, by row.
        if sort is "column", the second priority will be by column
        """
        if sort == "row": return [_range.rows_from_range("cell") for _range in self.ranges]
        elif sort == "column": return [_range.columns_from_range("cell") for _range in self.ranges]

class Coordinate():
    """ Represents a row and/or column index

    When both row and column are supplied, represents a specific cell index (e.g.- [Cell] A1).
    If the the first argument is a string and no other arguments are passed, the string should represent a Excel-format
    coordinate (either A1 or R1C1).
    If one of row or column is supplied and the other is None, represents a row or column index, respectively (i.e.- [Column] A).

    Row or column may individually be defined as a Excel-formatted string, an integer, or a tuple. Positive integers are always
    treated as absolute references where as strings are only treated as relative by default. A tuple should contain either
    the integer or string index and can have an optional second index indicating that it is an absolute reference (supplying
    either True, "$", or "absolute"). Row and column indexes can be negative only so long as they are not absolute (negative integers
    will be treated as relative).

    Each part of a Coordinate instance (row and column) is a namedtuple called Index with parts (type, value, absolute) where type indicates "row" or "column",
    value is the reference, and absolute is a boolean (True or False).
    """
    def __init__(self,row,column = False):
        if isinstance(row,str) and column is False:
            row,column = converttotuple(row)
        else:
            if column is False: column = None
            try: row,column = converttotuple((row,column))
            except Exception as e:
                raise AttributeError(f"Invalid values for cooridinate: {row},{column}")
        self._row = row
        self._column = column
    @property
    def row(self):
        return self._row.value if self._row else None
    @property
    def column(self):
        return self._column.value if self._column else None
    def isabsolute(self):
        return self._column.absolute or self._row.absolute
    def toA1string(self, absolute = True):
        """ Returns the Coordinate in A1 format.

            By default, will output absolute dollar signs (ex.- A$1).
            If absolute is False, will output both row and column as relative.
        """
        if not absolute:
            return f"{utils.cell.get_column_letter(self._column.value)}{self._row.value}"
        return f'{"$" if self._row.absolute else ""}{utils.cell.get_column_letter(self._column.value)}{"$" if self._column.absolute else ""}{self._row.value}'
    def __add__(self,other):
        if isinstance(other,Coordinate):
            if (self._row.absolute and other._row.absolute)\
               or self._column.absolute and other._column.absolute:
                raise AttributeError(f"Cannot add two Absolute references: {self} + {other}")
            return Coordinate(row = addindices(self._row,other._row), column = addindices(self._column, other._column))
        ## Otherwise, try to convert to Coordinate
        ## Conversion will automatically handle Address Strings and 2-length iterables (i.e. (row,column) Tuple)
        try: return self + Coordinate(other)
        except: pass
    def __eq__(self,other):
        if isinstance(other,Coordinate):
            return self._row == other._row and self._column == other._column
    def __repr__(self):
        return f"{self.__class__}({self._row},{self._column})"

Index = collections.namedtuple("indextuple","type, value, absolute")

## Order than Coordinates should be displayed
COORDINATEORDER = ["row","column"]

FULLRE = re.compile("""
^(?P<match>
    (?P<RC>
        (?P<RCrow>R                                                 ## R1C1
            (?P<absolute1>\[?)(?P<RCrowid>-?\d+)(?P<absolute2>\]?)
        )
        (?P<RCcolumn>C
            (?P<absolute3>\[)?(?P<RCcolumnid>-?\d+)(?P<absolute4>\]?)
        )
    )
    |
    (?P<A1>                                                         ## A1
        (?P<A1column>
            (?P<absoluteB>\$?)(?P<A1columnid>[A-Z]+)
        )
        (?P<A1row>
            (?P<absoluteA>\$?)(?P<A1rowid>\d+)
        )
    )                 
)$
""", re.VERBOSE | re.IGNORECASE)
COLUMNRE = re.compile("""
(?P<match>
      C(?P<absolute1>\[?)(?P<RCcolumn>-?\d+)(?P<absolute2>\]?)    ## R1C1
    | (?P<absolute>\$?)(?P<A1column>[A-Z]+)                     ## A1
)
""", re.VERBOSE | re.IGNORECASE)
ROWRE = re.compile("""
(?P<match>
      R(?P<absolute1>\[?)(?P<RCrow>-?\d+)(?P<absolute2>\]?)   ## R1C1
    | (?P<absolute>\$?)(?P<A1row>\d+)                       ## A1
)
""", re.VERBOSE | re.IGNORECASE)

def converttotuple(value):
    if isinstance(value,str):
        regex = FULLRE.search(value)
        if regex:
            syntax = "RC" if regex.group("RC") else "A1"
            row = parsecoordregex(regex.group(f"{syntax}row"),'row')
            column = parsecoordregex(regex.group(f"{syntax}column"),'column')
            return row,column
        else:
            raise TypeError(f"Invalid coordinates: {value}")
    try: r,c = value
    except: raise TypeError(f"Cooridinates must be Coordinate-Formatted string or a tuple: {value}")
    ## Using list so that we can freely swap the indices when necessary
    indices = [parseindex(r),parseindex(c)]
    ###       I'm handling the two indices ambiguously rather than simply calling one "row"      ###
    ###  and the other "column" because assuming so causes complexity in the code to compensate  ###
    ################################################################################################
    ### Dear Past Self:                                                                          ###
    ###     Examples, please...                                                                  ###
    ### Sincerely, Future Self                                                                   ###

    ## Handle Non-Indexes
    nonindices = [index.value for index in indices if index.value is None]
    if nonindices:
        return handlenonindices(indices)

    ## row/column are ambiguous (because they were both a numeric index instead of Excel notation)
    if all(index.type is None for index in indices):
        ## Assign their types in order
        indices[0] = indices[0]._replace(type = "row")
        indices[1] = indices[1]._replace(type = "column")
        return indices

    ## Indices are of the same type: the only way this should happen is because of the User
    ## e.g.- Coordinate("A","A"), Coordinate("1","1")
    if indices[0].type == indices[1].type:
        ## If both are rows, then we can try assume one to actually be a column index
        if indices[0].type == "row" and indices[1].type == "row":
            ## If only one is absolute, then the other can be assumed to be relative
            if indices[0].absolute and not indices[1].absolute:
                indices[1] = indices[1]._replace(type='column')
            elif indices[1].absolute and not indices[0].absolute:
                indices[1] = indices[0]._replace(type='column')
            elif not indices[0].absolute and not indices[1].absolute:
                indices[1] = indices[1]._replace(type='column')
            else:
                raise ValueError(f"Duplicated Arguments: {indices[0].type},{indices[1].type}")
        else:
            raise ValueError(f"Duplicated Arguments: {indices[0].type},{indices[1].type}")

    row = [index for index in indices if index.type == "row"]
    column = [index for index in indices if index.type == "column"]
    ambiguous = [index for index in indices if index.type is None]

    ## Both Ambiguous is handled above
    ## Ambiguous Index is Column
    if row and not column:
        row = row[0]
        column = ambiguous[0]
        column = column._replace(type = "column")
    ## Ambiguous Index is Row
    elif column and not row:
        column = column[0]
        row = ambiguous[0]
        row = row._replace(type = "row")
    ## No Ambiguous indices
    else:
        row = row[0]
        column = column[0]
    return row,column

def handlenonindices(indices):
    """ Fork to handle missing coordinate values """
    nonindices = [index.value for index in indices if index.value is None]
    ## Both are None indexes (p.s.- >= is extraneous, but we're doing it anyway)
    if len(nonindices) >= 2:
        raise ValueError(f"No Indicies provided: {indices}")

    goodind = [i for i,index in enumerate(indices) if index.value is not None][0]
    goodindex = indices[goodind]
    ## If goodindex is ambiguous, then we need to set it's type
    if goodindex.type is None:
        ## type is determined by position in the tuple 
        ## i.e. (None,1) -> (NoneIndex,AmbiguousIndex==1) -> (NoneRowIndex,ColumnIndex==1)
        goodindex = goodindex._replace(type = COORDINATEORDER[goodind])
    ## Otherwise, we need to double-check goodindex's position in the tuple
    else:
        ## If goodindex is not in the correct position
        if goodindex.type != COORDINATEORDER[goodind]:
            ## Swap the coordinates
            indices[0],indices[1] = indices[1],indices[0]
    ## By this point the indices should read (row,None) or (None,column)
    return indices

def parseindex(value):
    """ Function for finding information about a specific descriptor """
    ## Handle Row/Column References ("$A" where one one index is None)
    if value is None:
        return Index(None,None,False)
    ## Check if coord is an int
    ## As it turns out, bool is a subclass of int :-/
    if isinstance(value,int) and not isinstance(value,bool):
        return Index(None,value,True)
    if isinstance(value,str):
        ## Check against Regex
        try:
            return parsecoordregex(value,"column")
        except:
            return parsecoordregex(value,"row")
    ## Validate a Index
    ## Note, namedtuple is an instance of tuple, so this has to happen before handling tuple/lists
    if isinstance(value,Index):
        ## Index should be "row","column" (, or None for Column References)
        if value.type not in ["row","column",None]: raise ValueError(f"Index's type should be 'row','column', or None: {value}")
        ## Index made by this module use column index (or None for Column References)
        if not isinstance(value.value,int) and value.value is not None: raise ValueError(f"Index's value is not an integer (or None): {value}")
        ## This module only uses True/False for absolute
        if value.absolute not in [True,False]: raise ValueError(f"Index's absolute value is not True or False: {value}")
        ## Index is valid
        return value
    ## Handle Tuples and Lists (possibly other iterables later)
    if isinstance(value,(tuple,list)):
        ## Only accept lengths 1,2
        if 0 > len(value) or len(value) > 2:
            raise ValueError(f"Coordinate part's length should be 1 or 2 (index,[absolute]): received length {len(value)}")
        index = parseindex(value[0])
        if len(value) == 2:
            absolute = value[1]
            ## Convert string
            if isinstance(absolute,str):
                absolute = absolute.lower()
                ## Acceptable absolute = True values
                if absolute in ["$","absolute"]: absolute = True
                ## Only absolute = False value
                elif absolute == "": absolute = False
                else: raise ValueError(f"Coordinate part's second index is an unknown string: {value[1]}")
            ## Acceptable values
            if absolute not in [0,1,True,False]:
                raise ValueError(f"Coordinate part's second index must be True or False, or an accepted alias: {absolute}")
            ## Check collision between index defined in value[0]:
            ## ("$A",False) seems to be the only relevant case
            if index.absolute and not absolute:
                raise ValueError(f"Coordinate part's second index contradicts the first: {value}")
            index = Index(index.type, index.value, absolute)
        return index

    ## Everything else
    raise ValueError(f"Coordinate part is not a recognized format: {value}")

def parsecoordregex(value,colrow):
    """ Ambidextrous helper function for parseindex """
    if colrow not in ("column","row"): raise ValueError(f"parsecoordregex must be 'column' or 'row': {colrow}")
    if colrow == "column": regex = COLUMNRE.search(value)
    else: regex = ROWRE.search(value)
    if not regex: raise ValueError(f"String does not match any identifiable {colrow} patterns: {value}")
    if regex.group(f"A1{colrow}"):
        index = regex.group(f"A1{colrow}")
        if colrow == "column": index = utils.cell.column_index_from_string(index)
        index = int(index)
        return Index(colrow, index, bool(regex.group("absolute")))
    else:
        abs1,abs2 = bool(regex.group("absolute1")),bool(regex.group("absolute2"))
        ## If both are not true, then we are missing one of the brackets
        if abs1 != abs2:
            ## If abs1 is True, then we are missing the closing bracket (abs2)
            missing = "]" if abs1 else "["
            raise SyntaxError(f"Incomplete coordinate description: missing {missing}.")
        return Index(colrow, int(regex.group(f"RC{colrow}")),abs1 and abs2)

def addindices(index1,index2):
    """ Adds two indices together. Both must have the same type and at least one must be relative (Index.absolute == False)"""
    if index1.absolute and index2.absolute:
        raise AttributeError(f"Cannot add two Absolute Indices: {index1} + {index2}")
    if index1.type != index2.type:
        raise TypeError(f"Type mismatch between Indices: {index1}, {index2}")
    ## Sort indices so that any absolute index is first
    ## (so only need to check if the second needs to be moved)
    if index2.absolute: index1,index2 = index2,index1
    ## If index1 is absolute, then the result should be absolute
    result = Index(index1.type, index1.value + index2.value, index1.absolute)
    ## If result is absolute, make sure that it is is still above 0
    ## (doesn't matter for relative Indexes)
    ## There doesn't seem to be any valid reason to use min(1), since the
    ## first index is always a known quantity and any user who wants to reduce
    ## a Coordinate/Index to the first index should just do so manually.
    if result.value <= 0:
        raise ValueError(f"Absolute Reference reduced below 0: {result}")
    return result

class Range():
    def __init__(self,worksheet,range):
        """ Creates a new Range Object

        worksheet is an openpyxl worksheet object
        range is a string expressing a valid, single range
        """
        self.worksheet = worksheet
        self._startcoord, self._endcoord = None,None
        self._masterstart, self._masterend = None,None
        coords = range.split(":")
        if 0 > len(coords) > 2:
            raise AttributeError(f"Invalid Range: {range}")
        self._startcoord = Coordinate(coords[0])
        if len(coords) == 2: self._endcoord = Coordinate(coords[1])
        else: self._endcoord = self.startcoord
        self.updatemasters()

    @property
    def startcoord(self):
        return self._startcoord
    @startcoord.setter
    def startcoord(self,value):
        if not isinstance(value,Coordinate):
            value = Coordinate(value)
        self._startcoord = value
        self.updatemasters()
    @property
    def endcoord(self):
        return self._endcoord
    @endcoord.setter
    def endcoord(self,value):
        if not isinstance(value,Coordinate):
            value = Coordinate(value)
        self._endcoord = value
        self.updatemasters()

    @property
    def masterstart(self):
        return self._masterstart
    @property
    def masterend(self):
        return self._masterend
    def updatemasters(self):
        c_1,r_1,c_2,r_2 = utils.cell.range_boundaries(self.range)
        self._masterstart = Coordinate(r_1,c_1)
        self._masterend = Coordinate(r_2,c_2)

    @property
    def range(self):
        return f"{self.startcoord.toA1string()}:{self.endcoord.toA1string()}"

    def rows_from_range(self,attribute = "value"):
        """ Returns cell property per openpyxl.utils.cell.rows_from_range
        
        If attribute is "address," functions as openpyxl.utils.cell.rows_from_range
        If attribute is "cell," returns cell references instead.
        """
        if attribute == "address":
            return utils.cell.rows_from_range(self.range)
        if attribute == "cell":
            return (map(lambda cell: self.worksheet[cell], row) for row in utils.cell.rows_from_range(self.range))
        return (map(lambda cell: getattr(self.worksheet[cell],attribute), row) for row in utils.cell.rows_from_range(self.range))

    def columns_from_range(self):
        """ Returns the cell values for the result of utils.cell.cols_from_range using self.sheet """
        return (map(lambda cell: self.worksheet[cell].value, column) for column in utils.cell.cols_from_range(self.range))

    def cells_by_row(self,attribute = "value"):
        """ Returns a flat list of cells based on rows_from_range
        
        attribute is the same as rows_from_range
        """
        return (cell for row in list(self.rows_from_range(attribute=attribute)) for cell in row)

    def cells_by_column(self,attribute = "value"):
        """ Returns a flat list of cells based on columns_from_range
        
        attribute is the same as rows_from_range
        """
        return (cell for column in list(self.columns_from_range(attribute=attribute)) for cell in column)

    def row(self,index,attribute = "value"):
        """ Returns the row with the provided zero-index, via rows_from_range """
        return list(self.rows_from_range(attribute=attribute))[index]

    def column(self,index,attribute = "value"):
        """ Returns the column with the provided zero-index, via columns_from_range """
        return list(self.columns_from_range(attribute=attribute))[index]

    def isinrange(self,cell):
        """ Returns whether the cell is in the Range """
        coord = Coordinate(cell.coordinate)
        return (self.masterstart.column <= coord.column <= self.masterend.column)\
            and (self.masterstart.row <= coord.row <= self.masterend.row)

    def coordinatetocell(self,coordinate):
        """ Converts a Coordinate (which may be relative) to a Cell (which is absolute) based on the first index of the range """
        ## Note that range_boundaries is a CR pattern rather than RC
        if not coordinate.isabsolute():
            if not coordinate._row.absolute:
                if coordinate.row >= 0:
                    row  = self.masterstart.row + coordinate.row
                else:
                    row = self.masterend.row + coordinate.row + 1
                coordinate._row = Index("row",row,True)
            if not coordinate._column.absolute:
                if coordinate.column >= 0:
                    column = self.masterstart.column + coordinate.column
                else:
                    column = self.masterend.column + coordinate.column + 1
                coordinate._column = Index("column",column,True)
        return self.worksheet.cell(row = coordinate.row, column = coordinate.column)

    def subrange(self,start = None, end = None):
        """ Returns a new Range Object that is a slice of this Range Object

        Takes both start and end. start and end should be appropriate coordinates:
        an A1/R1C1 Notation string, tuple of (row,column) indexes, or Coordinate Object.
        Tuples follow the same syntax as Coordinate objects and must be length two.

        For tuples None can be substituted for either the value itself or either of the
        tuple indexes; as a replacement for start or end, it refers to the address of the
        (first-column,first-row) or the (last-column,last-row) respectively. As part
        of the tuple it refers to that part of the default value, i.e.- None for
        start.row is an alias for first-row, end.column = None references last-column, etc.

        Example - Range(C4:J6).subrange(end=(1,None)) == Range(C4:F4)

        String Notation and 
        """

        ## Note that range_boundaries is a CR pattern rather than RC
        if start is None: start = [(self.masterstart.row,True),(self.masterstart.column,True)]
        if isinstance(start,(tuple,list)):
            if len(start) != 2: raise ValueError(f"Start and end tuples must be length 2: {start}")
            start = list(start)
            if start[0] is None: start[0] = (self.masterstart.row,True)
            if start[1] is None: start[1] = (self.masterstart.column,True)
            start = Coordinate(*start)
        else: start = Coordinate(start)
        start = self.coordinatetocell(start)
        
        if end is None: end = [(self.masterend.row,True),(self.masterend.column,True)]
        if isinstance(end,(tuple,list)):
            if len(end) != 2: raise ValueError(f"Start and end tuples must be length 2: {end}")
            end = list(end)
            if end[0] is None: end[0] = (self.masterend.row,True)
            if end[1] is None: end[1] = (self.masterend.column,True)
            end = Coordinate(*end)
        else: end = Coordinate(end)
        end = self.coordinatetocell(end)

        
        if not self.isinrange(start) or not self.isinrange(end):
            start = Coordinate(start.row,start.column)
            end = Coordinate(end.row, end.column)
            raise ValueError(f"Subrange coordinates out of Range: R[{self.masterstart.row}]C[{self.masterstart.column}] : R[{self.masterend.row}]C[{self.masterend.column}] <> R[{start.row}]C[{start.column}] : R[{end.row}]C[{end.column}]")

        return Range(self.worksheet,f"{start.column_letter}{start.row}:{end.column_letter}{end.row}")
        
    def __eq__(self,other):
        if isinstance(other,Range):
            return self.worksheet == other.worksheet and self.startcoord == other.startcoord and self.endcoord == other.endcoord

    def __repr__(self):
        return f"{self.__class__.__name__}({self.range})"

"""
NOTES ON TABLE ATTRIBUTES
(Since the source Docs aren't terribly useful)


"""

def gettablesize(sheet,startcolumn,startrow, absolute = False):
    """ This is a function to intuit the shape of a data series which is laid out in a table format.

        It works by assuming the top row to be the header row. It scans this row until it reaches a blank cell.
        It then iterates over the subsequent rows up to the end of the header row. When a row is completely
        blank it breaks and returns a rage from the first cell of the header row to the last row and last column
        of the header row.

        The absolute keyword determines whether the returned range is absolute (default False).

        e.x.- Table Starting at (1,1) [A1]
        ---------------
        |W   X   Y   Z|
        |1   2   3    |
        |    4       5|
        |6            |
        |             |
        |[... etc ...]|
        ---------------
        Returns a range "A$1$:D$4$"
    """
    row,column = startrow,startcolumn
    cell = sheet.cell(row = row, column = column)
    while cell.value:
        column += 1
        cell = sheet.cell(row = row, column = column)

    endcolumn = column - 1 
    def checkrow(row):
        col = 1
        while col <= endcolumn:
            if sheet.cell(row = row, column = col).value:
                return True
            col += 1
    ## Start check table data 
    row = 2 
    rowhasvalue = checkrow(row)
    while rowhasvalue:
        row += 1
        rowhasvalue = checkrow(row)
    ## Current Row does not have value
    endrow = row - 1
    return tuple_to_range(startcolumn,startrow,endcolumn,endrow, absolute = absolute)

def tuple_to_range(*trange, absolute = False):
    """ Returns an Excel-version (A1 notation) string representing the given range.
        
        The tuple should be 1-indexed, can be passed as a single length-4 tuple or as 4 positional arguments,
        and should represent (startcolumn, startrow, endcolumn, endrow).

        The absolute keyword determines whether the returned range is absolute (default False).
    """
    if len(trange) == 1:
        trange = trange[0]
    if len(trange) != 4:
        raise ValueError("tuple range should be length 4.")
    c1,r1,c2,r2 = trange
    if absolute:
        c1,r1,c2,r2 = int(c1),int(r1),int(c2),int(r2)
    else:
        c1,r1,c2,r2 = str(c1),str(r1),str(c2),str(r2)
        
    start,end = Coordinate(r1,c1),Coordinate(r2,c2)
    return f"{start.toA1string()}:{end.toA1string()}"


class EnhancedTable(openpyxl.worksheet.table.Table):
    """ A better Table Class, returned by get_all_tables """
    def from_table(table,worksheet):
        ## NOTE! It seems that Initializing a Table object (which EnhancedTable is a subclass of) automatically adds the table to the Worksheet's _tables list
        ## This means that we have to check and remove the original Table object, and should return any pre-built EnhancedTable we find.
        oldtables = [tab for tab in worksheet._tables if tab == table]
        for tab in oldtables:
            ## Don't bother making a new one
            if isinstance(tab,EnhancedTable): return tab
            ## Remove the old-style one
            else: worksheet._tables.remove(tab)
        ## If no EnhancedTable version found, create one (which- again- seems to automatically be added to the _tables list)
        return EnhancedTable(worksheet = worksheet,
                             id=table.id, displayName=table.displayName, ref=table.ref, name=table.name, comment=table.comment, tableType=table.tableType,
                             headerRowCount=table.headerRowCount, insertRow=table.insertRow, insertRowShift=table.insertRowShift, totalsRowCount=table.totalsRowCount, totalsRowShown=table.totalsRowShown,
                             published=table.published, headerRowDxfId=table.headerRowDxfId, dataDxfId=table.dataDxfId, totalsRowDxfId=table.totalsRowDxfId, headerRowBorderDxfId=table.headerRowBorderDxfId,
                             tableBorderDxfId=table.tableBorderDxfId, totalsRowBorderDxfId=table.totalsRowBorderDxfId, headerRowCellStyle=table.headerRowCellStyle, dataCellStyle=table.dataCellStyle,
                             totalsRowCellStyle=table.totalsRowCellStyle, connectionId=table.connectionId, autoFilter=table.autoFilter, sortState=table.sortState, tableColumns=table.tableColumns,
                             tableStyleInfo=table.tableStyleInfo, extLst=table.extLst)
    def __init__(self,worksheet, **kw):
        self.worksheet = None
        super().__init__(**kw)
        self.worksheet = worksheet
        self.range = Range(worksheet,self.ref)
        ## For the moment, we are disabling this functionality
        ## The serializer on Table seems to be serializing other attributes
        ## self.worksheet.add_table(self)

    @property
    def ref(self):
        return self.range.range
    @ref.setter
    def ref(self,value):
        self.range = Range(self.worksheet,value)

    def insertcolumn(self,index = None,columnname = None):
        collen = len(self.tableColumns)
        if index is None:
            index = collen - 1
        if index >= collen: raise ValueError("Index outside of Table Range")
        if columnname is None:
            columnname = f"Column{collen}"
        if columname in [column.name for column in self.tableColumns]:
            columnname = columnname + " 2"
        self.tableColumns.insert(index,openpyxl.worksheet.table.TableColumn(id=collen,name=columnname))
        self.range += [1,None]

    def headers(self, attribute = "value"):
        """ Returns the headers of the EnhancedTable as a list, via self.range.row """
        ## Last row (equal to row count) should contain actual headers... In theory
        ## And we zero-index it
        headerrow = self.headerRowCount - 1
        return list(self.range.row(headerrow,attribute=attribute))

    def addheaders(self, headers):
        """ Adds a set of headers to the table at the bottom of the header range """
        headerlen = len(headers)               

    def headerrange(self):
        """ Returns the Table's headers as a Range object """
        headerlength = self.headerRowCount
        columnlength = len(self.tableColumns)
        return self.range.subrange(None,(str(headerlength-1),str(columnlength-1)))

    def datarange(self):
        """ Returns the Table's body as a Range object """
        headerlength = self.headerRowCount
        columnlength = len(self.tableColumns)
        return self.range.subrange(
            (str(headerlength),str(0)),None)

    def todicts(self,keyfactory = None):
        """ Converts all data rows to dicts based on column headers. The first element of the returned list is a list of the header strings used.
        
        keyfactory is an callback function to modify the keys (example- the lowerstrip lambda available in this module executes
        key.lower().replace(" ","_") to lowercase and remove spaces in keys).
        """
        if keyfactory is None: keyfactory = lambda key: key
        headers = [keyfactory(key) for key in self.headers()]

        data = [collections.OrderedDict(list(zip(headers,row))) for row in self.datarange().rows_from_range()]
        data.insert(0,headers)
        return data

def get_all_tables(workbook):
    """ Returns a list of tuples of all tables in the workbook formatted: (worksheetobject, EnhancedTable Object) """
    out = []
    for worksheetname in workbook.sheetnames:
        worksheet = workbook[worksheetname]
        ## To ensure list integrity, we'll have to copy the list
        ## (Initiating Tables seems to automatically add them to _tables list)
        tables = list(workbook[worksheetname]._tables)
        for table in tables:
            ## Check if it was pre-converted and return it if so
            if isinstance(table,EnhancedTable):
                out.append((worksheet,table))
            ## Otherwise, convert it ourselves
            else:
                out.append((worksheet,EnhancedTable.from_table(table,worksheet)))
    return out

def get_table_by_name(worksheet,name):
    """ Returns the table with the given displayName.

        Workbooks that have duplicate tables are considered Invalid by Excel,
        so if this method finds multiple tables with the given displayName it
        will raise a ValueError.
    """
    if not isinstance(worksheet,openpyxl.worksheet.worksheet.Worksheet):
        raise TypeError("worksheet should be a Worksheet object")
    if not isinstance(name,str):
        raise TypeError("name should be a string")
    for table in worksheet._tables:
        if table.displayName == name:
            return table

def dicts_to_table(sheet, dinput, tablename = None, start = None, headers = None):
    """ Writes a list of dictionaries or lists into a Table.
  
        sheet should be a Worksheet.
        dinput should be a list of either lists or dictionaries.
        If provided, tablename should be a string. If the tablename is defined
        within the worksheet, then all rows in the given table will be replaced
        with dinput.
        If provided, start should be a valid coordinate to place the table. It is
        a SyntaxError to supply both an existing tablename and start and a ValueError
        to supply neither.
        headers should be a list of strings. headers is optional if dinput is a
        list of dicts. If so, then only the given keys will be outputted from dinput.
        If dinput is a list and tablename is not an existing table, then headers is
        required. Otherwise (dinput is a list and tablename exists), then headers is
        ignored.
    """
    if not isinstance(sheet,openpyxl.worksheet.worksheet.Worksheet):
        raise TypeError('sheet must be a worksheet')
    if not isinstance(dinput,(list,tuple)) or not all(isinstance(item,(dict,list,tuple)) for item in dinput):
        raise TypeError('dinput should be a list of lists or dicts')
    t1 = dinput[0].__class__
    if any(not isinstance(item,t1) for item in dinput):
        raise TypeError("All dinputs must be the same class")
    if tablename is None: tablename = ""
    if not isinstance(tablename,str):
        raise TypeError('If supplied, tablename should be a string')
    table = None
    if tablename:
        table = gettablebyname(sheet,tablename)
    else:
        ## AutoGenerate Tablename
        i = 1
        while not tablename:
            tablename = f"Table{i}"
            t = get_table_by_name(sheet,tablename)
            if t: tablename = None

    if start:
        start = Coordinate(start)
        if table and table.startcoord != start:
            raise SyntaxError("Cannot supply both an existing table and start")
    elif table:
        start = table.startcoord
    ## Not start and not table
    else:
        raise ValueError("dicts_to_table requires either start or an existing tablename")

    if headers:
        if not isinstance(headers,(list,tuple)) or any(not isinstance(item,str) for item in headers):
            raise ValueError("headers should be a string")
        ## Ignore headers if existing table
        if table: headers = None

    ## Normalize Output
    if headers is None:
        if table:
            headers = table.headers()

    ## Normalize dinput structure
    if issubclass(t1,dict):
        if headers is None:
            iheaders = list(itertools.chain.from_iterable([item.keys() for item in dinput]))
            ## Instead of using set, we're going to attempt to conserve at least some of the order of the keys
            ## (at time of writing, all standard dicts memorize input order of keys, which makes them reliable)
            headers = []
            for h in iheaders:
                if h not in headers: headers.append(h)
        dinput = [[item.get(key,"") for key in headers] for item in dinput]
    ## Otherwise, dinput items are lists or tuples (or- theoretically- subclasses), so don't do anything
    
    ## If existing table, remove it and rewrite it
    if table:
        for cell in table.range.cells_by_row(value = "cell"):
            cell.value = None
        table.worksheet._tables.remove(table)
    ## TODO: we should probably make some attempt to NOT overwrite data in proximity if table extends beyond original constraints

    ## Write Table Headers
    for coffset,column in enumerate(headers):
        offset = Coordinate(row=str(0),column=str(coffset))
        cell = start + offset
        cell = sheet[cell.toA1string()]
        cell.value = str(column)
    ## Write table data
    ## Start writing row at 1
    for roffset,row in enumerate(dinput, start = 1):
        for coffset,column in enumerate(headers):
            offset = Coordinate(str(roffset),str(coffset))
            cell = start + offset
            cell = sheet[cell.toA1string()]
            cell.value = row[coffset]
    ## Add/Re-add Table to sheet
    endcell = Coordinate(str(roffset),str(coffset))
    endcell = start + endcell
    table = openpyxl.worksheet.table.Table(displayName = tablename, ref = f"{start.toA1string(False)}:{endcell.toA1string(False)}")
    sheet.add_table(table)
    table = EnhancedTable.from_table(table,sheet)
    return table
